#!/usr/bin/env python
#  -*- coding: utf-8 -*-
"""
@File    :   eval_reponses.py
@Time    :   2024/03/09 09:26:01
@Version :   1.0
@Desc    :   recuperation des réponses "google form" et excel
"""
import copy
import os
import datetime
import pprint
import pickle
import subprocess
import json
import shutil

# import codecs
# import zipfile
# import csv

from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
#from google.auth.transport.requests import Request

from gforms import Form

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.properties import PageSetupProperties
#from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def print_form(form_id):
    url = f'https://docs.google.com/forms/d/{form_id}/edit'
    form = Form()

    form.load(url)
    print(form.to_str(indent=2))  # a text representation, may be useful for CLI applications


def retrieve_from_google_forms(form_id):
    SCOPES = "https://www.googleapis.com/auth/forms.body.readonly https://www.googleapis.com/auth/forms.responses.readonly"
    credentials = None
    if os.path.exists('credentials.dat'):
        with open('credentials.dat', 'rb') as credentials_dat:
            credentials = pickle.load(credentials_dat)
        if credentials.expired:
            os.remove('credentials.dat')
    if not os.path.exists('credentials.dat'):
        flow = InstalledAppFlow.from_client_secrets_file(
            'client_secret.json', SCOPES)
        credentials = flow.run_local_server()
        with open('credentials.dat', 'wb') as credentials_dat:
            pickle.dump(credentials, credentials_dat)
    service = build('forms', 'v1', credentials=credentials)

    _questions = service.forms().get(formId=form_id).execute()
    _reponses = service.forms().responses().list(formId=form_id).execute()
    if 'responses' in _reponses.keys():
        return _questions['items'], _reponses['responses']
    return _questions['items'], []


def iter_dict(dico, item2find):
    if isinstance(dico, (tuple, list, set, frozenset)):
        for _dico in dico:
            iter_dict(_dico, item2find)
    else:
        for clef, valeur in dico.items():
            if clef == item2find:
                print(valeur)
                pprint.pp(dico)
            elif isinstance(valeur, (tuple, list, set, frozenset, dict)):
                iter_dict(valeur, item2find)
            elif 'title' in clef.lower():
                print(valeur)


def vers_xlsx(dico, nom, xlsx, _freeze=None):
    xlsx.create_sheet(nom)
    page = xlsx[nom]
    indice1 = 1
    indice2 = 1
    for clef1, valeur1 in dico.items():
        if isinstance(valeur1, (dict)):
            cellule = page.cell(row=1, column=indice1) 
            cellule.value = clef1
            indice1 += len(valeur1.keys())
            for clef2, valeur2 in valeur1.items():
                cellule = page.cell(row=2, column=indice2) 
                cellule.value = clef2
                debut = 3
                for valeur3 in valeur2:
                    cellule = page.cell(row=debut, column=indice2) 
                    cellule.value = valeur3
                    debut += 1
                indice2 += 1
        elif isinstance(valeur1, (list)):
            cellule = page.cell(row=1, column=indice1) 
            cellule.value = clef1
            debut = 2
            for valeur2 in valeur1:
                cellule = page.cell(row=debut, column=indice1) 
                cellule.value = valeur2
                debut += 1
            indice1 += 1
        else:
            page.append([clef1, valeur1])
    #
    page.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    for col in page.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            if cell.value:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
        adjusted_width = (max_length + 1) * 1.005
        if adjusted_width > 80:
            adjusted_width = 80.0
        page.column_dimensions[column].width = adjusted_width
    if _freeze:
        page.freeze_panes = _freeze


def comparaison_xslx(fichier1, fichier2):
    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    wk1 = load_workbook(fichier1)
    wk2 = load_workbook(fichier2)
    compare = 'ligne'
    for page in wk1.sheetnames:
        if wk1[page].freeze_panes is not None:
            if wk1[page].freeze_panes == 'B2':
                compare = 'ligne'
            else:
                compare = 'colone'
        if compare == 'ligne':
            lignes_a_comparer = list(wk2[page].iter_rows(values_only=True))
            for ligne in wk1[page].iter_rows():
                valeurs = tuple([cell.value for cell in ligne])
                if valeurs not in lignes_a_comparer:
                    for cellule in ligne:
                        cellule.fill = redFill
        if compare == 'colone':
            max_col = max(wk1[page].max_column, wk2[page].max_column)
            for i in range(max_col):
                try:
                    if len(wk2[page][get_column_letter(i)]) > len(wk1[page][get_column_letter(i)]):
                        wk1[page][get_column_letter(i)][1].fill = redFill
                    col_a_comparer = [cell.value for cell in wk2[page][get_column_letter(i)]]
                    for cellule in wk1[page][get_column_letter(i)]:
                        if cellule.value not in col_a_comparer:
                            cellule.fill = redFill
                except ValueError:
                    if i == 0:
                        pass
    wk1.save(fichier1)
    pass


if __name__ == '__main__':
    debug = False
    # from_zip()
    questions = []
    reponses = []
    
    with open("forms_id.json", "r") as fic:
        forms = json.load(fic)
        
    if debug:
        for form_id, _dat in forms.items():
            print_form(form_id)
    
    for form_id, _dat in forms.items():
        if os.path.exists(_dat):
            # vérification de la date d'export
            now = datetime.datetime.now()
            debut = now - datetime.timedelta(days=1)
            st = os.stat(_dat)
            date_f = datetime.datetime.fromtimestamp(st.st_mtime)
            if date_f < debut:
                os.remove(_dat)
        if not os.path.exists(_dat):
            _questions, _reponses = retrieve_from_google_forms(form_id)
            with open(_dat, 'wb') as responses_dat:
                pickle.dump([_questions, _reponses], responses_dat)
        else:
            with open(_dat, 'rb') as responses_dat:
                _questions, _reponses = pickle.load(responses_dat)
        for item in _questions:
            questions.append(item)
        for item in _reponses:
            reponses.append(item)

    # Mise en format du dictionnaire des questions
    if debug:
        pprint.pp(questions)
        with open("questions.json", "w") as fic:
            fic.write(pprint.pformat(questions))

    question_dico = {}
    choice_dico = {}
    for item_dict in questions:
        if 'questionItem' in item_dict.keys():
            if 'choiceQuestion' in item_dict['questionItem']['question'].keys():
                if item_dict['itemId'] not in choice_dico.keys():
                    choice_dico[item_dict['itemId']] = {'choix': item_dict['title'], 'valeurs': []}
                _list_valeurs = []
                for valeur in item_dict['questionItem']['question']['choiceQuestion']['options']:
                    try:
                        _list_valeurs.append(valeur['value'])
                    except KeyError:
                        for _clef in valeur.keys():
                            _list_valeurs.append(_clef)
                            break
                choice_dico[item_dict['itemId']]['valeurs'].append(_list_valeurs)
            if item_dict['questionItem']['question']['questionId'] not in question_dico.keys():
                question_dico[item_dict['questionItem']['question']['questionId']] = []
            question_dico[item_dict['questionItem']['question']['questionId']].append(item_dict['title'])
        elif 'questionGroupItem' in item_dict.keys():
            if item_dict['itemId'] not in choice_dico.keys():
                choice_dico[item_dict['itemId']] = {'choix': item_dict['title'], 'valeurs': []}
            _list_valeurs = []
            for valeur in item_dict['questionGroupItem']['grid']['columns']['options']:
                _list_valeurs.append(valeur['value'])
            choice_dico[item_dict['itemId']]['valeurs'].append(_list_valeurs)
            for _quest in item_dict['questionGroupItem']['questions']:
                if _quest['questionId'] not in question_dico.keys():
                    question_dico[_quest['questionId']] = []
                question_dico[_quest['questionId']].append("%s %s" % (
                    item_dict['title'], ''.join(_quest['rowQuestion']['title'])))
    if debug:
        pprint.pp(question_dico)
    if debug:
        pprint.pp(reponses)
    dico_choix = {}
    for clef in choice_dico.keys():
        if len(choice_dico[clef]['valeurs'][0]) * 2 != len(choice_dico[clef]['valeurs'][1]) + len(choice_dico[clef]['valeurs'][2]):
            # incohérence => on ignore
            continue
        dico_choix[choice_dico[clef]['choix']] = []
        for ind2 in range(len(choice_dico[clef]['valeurs'][0])):
            dico_choix[choice_dico[clef]['choix']].append([])
            for ind1 in range(len(choice_dico[clef]['valeurs'])):
                dico_choix[choice_dico[clef]['choix']][ind2].append(choice_dico[clef]['valeurs'][ind1][ind2])
    if debug:
        pprint.pp(dico_choix)
    liste_reponses = []
    for item_dict in reponses:
        resp_dict = {'courriel': item_dict['respondentEmail']}
        for _clef, _item in item_dict['answers'].items():
            resp_dict[question_dico[_clef][0]] = []
            for valeur in _item['textAnswers']['answers']:
                resp_dict[question_dico[_clef][0]].append(valeur['value'])
        liste_reponses.append(copy.copy(resp_dict))
    
    # traitement des réponses
    if debug:
        for _id, valeur in question_dico.items():
            print("%s : %s" % (_id, valeur))

    liste_inscrits = {
        'Nom': [],
        'Téléphone': [],
        'Prieuré': [],
        'Région': [],
        'Chapitre': [],
        'samedi matin': [],
        'samedi après-midi': [],
        'dimanche matin': [],
        'dimanche après-midi': [],
        'lundi matin': [],
        'langues': []
    }
    liste_colones = {
        'Je marche sur la colonne adulte ou enfant ? adultes': 'Adultes',
        'Je marche sur la colonne adulte ou enfant ? enfants': 'Enfants',
        'Je marche sur la colonne adulte ou enfant ? adolescents': 'Adolescents',
        'Je marche sur la colonne adulte ou enfant ? je ne marche pas': ''
    }
    arrivee_differente = {}
    depart_different = {}
    transport = {}
    repas_adulte = {
        'samedi - petit-déjeuner': 0,
        'samedi - déjeuner (pique-nique)': -1,  # correctif mauvaise inscription
        'samedi - dîner (à table)': 0,
        'dimanche - petit-déjeuner': 0,
        'dimanche - déjeuner (pique-nique)': -1,  # correctif mauvaise inscription
        'dimanche - apéritif et dîner avec tous les confrères présents (à table)': 0,
        'lundi - petit-déjeuner': 0,
        'lundi - déjeuner (pique-nique)': 0,
        'lundi - dîner (pique-nique à emporter)': 0
    }
    repas_adulte_nominatif = {
        'samedi - petit-déjeuner': [],
        'samedi - déjeuner (pique-nique)': [],
        'samedi - dîner (à table)': [],
        'dimanche - petit-déjeuner': [],
        'dimanche - déjeuner (pique-nique)': [],
        'dimanche - apéritif et dîner avec tous les confrères présents (à table)': [],
        'lundi - petit-déjeuner': [],
        'lundi - déjeuner (pique-nique)': [],
        'lundi - dîner (pique-nique à emporter)': []
    }
    repas_enfant_nominatif = {
        'samedi - déjeuner (pique-nique)': [],
        'dimanche - déjeuner (pique-nique)': [],
    }
    repas_enfant = {
        'samedi - déjeuner (pique-nique)': -2,  # correctif mauvaise inscription
        'dimanche - déjeuner (pique-nique)': -1,  # correctif mauvaise inscription
    }
    porteurs_de_croix = {
        'samedi matin': [],
        'samedi après-midi': [],
        'dimanche matin': [],
        'dimanche après-midi': [],
        'lundi matin': []
    }
    regulation_securite = {
        'samedi matin': [],
        'samedi après-midi': [],
        'dimanche matin': [],
        'dimanche après-midi': [],
        'lundi matin': []
    }
    logistique_bivouacs = {
        'samedi matin': [],
        'samedi après-midi': [],
        'dimanche matin': [],
        'dimanche après-midi': [],
        'lundi matin': []
    }
    logistique_haltes = {
        'samedi matin': [],
        'samedi après-midi': [],
        'dimanche matin': [],
        'dimanche après-midi': [],
        'lundi matin': []
    }
    eclopes = {
        'samedi matin': [],
        'samedi après-midi': [],
        'dimanche matin': [],
        'dimanche après-midi': [],
        'lundi matin': []
    }
    aides_chapitres = {
        'samedi matin': [],
        'samedi après-midi': [],
        'dimanche matin': [],
        'dimanche après-midi': [],
        'lundi matin': []
    }
    messes_greffiers = {
        'samedi': [],
        'dimanche': [],
        'lundi': []
    }
    messes_trou_moreau = {
        'samedi': [],
        'dimanche': [],
        'lundi': []
    }
    messes_st_nic = {
        'samedi': [],
        'dimanche': [],
        'lundi': []
    }
    messes_chartres = {
        'samedi': [],
        'dimanche': [],
        'lundi': []
    }
    messes_logistiques = {
        "samedi 18h30": [],
        "samedi 20h00": [],
        "dimanche 12h30": [],
        "dimanche 16h30": [],
        "dimanche 18h30": [],
        "dimanche 19h30": [],
    }
    divers = {
        'lit en tente prêtres': 0,
        ' ': ' ',
        'Observations ou souhaits particuliers': ' ',
        '  ': ' ',
    }
    dico_nom_contacts = {}
    for item_dict in liste_reponses:
        if 'Nom' not in item_dict.keys() and 'Name' not in item_dict.keys():
            if debug:
                print('%s ne participera pas au pèlerinage' % item_dict['courriel'])
            continue
        try:
            nom = '%s %s %s' % (item_dict['Titre'][0], item_dict['Nom'][0].upper(), item_dict['Prénom'][0])
        except KeyError:
            nom = '%s %s' % (item_dict['Titre'][0], item_dict['Nom'][0].upper())
        dico_nom_contacts[nom] = item_dict['courriel']
        liste_inscrits['Nom'].append(nom)
        try:
            liste_inscrits['Prieuré'].append(item_dict['Prieuré, couvent ou paroisse (nom et localité)'][0])
        except KeyError:
            liste_inscrits['Prieuré'].append('')
        try:
            liste_inscrits['Région'].append(item_dict['Région :'][0])
        except KeyError:
            liste_inscrits['Région'].append('')
        try:
            liste_inscrits['Chapitre'].append(item_dict['Chapitre :'][0])
        except KeyError:
            liste_inscrits['Chapitre'].append('')
        try:
            liste_inscrits['Téléphone'].append(item_dict['Téléphone mobile'][0])
        except KeyError:
            liste_inscrits['Téléphone'].append('')
        try:
            liste_inscrits['langues'].append(', '.join(item_dict['Langues proposées pour les confessions']))
        except KeyError:
            liste_inscrits['langues'].append('')
        for demis in dico_choix['Je marche sur la colonne adulte ou enfant ?']:
            presence_demi = ''
            for colone in liste_colones.keys():
                if colone in item_dict.keys():
                    for demi in item_dict[colone]:
                        if demi in demis:
                            presence_demi = liste_colones[colone]
            liste_inscrits[demis[0]].append(presence_demi)
        if debug:
            print('- %s:' % nom)
        for _clef, _item in item_dict.items():
            if _clef in ['Nom et Prénom', 'Titre', 'Téléphone mobile', 'courriel']:
                continue
            if debug:
                print("    %s: %s" % (_clef, _item))
            # Arrivée en cours de route
            if  'Si j\'arrive en cours de route, je précise quand :' in _clef:
                _jour = _clef.split(' : ')[1]
                if item_dict[_clef][0] in dico_choix['Si j\'arrive en cours de route, je précise quand :'][0]:
                    continue
                if _jour not in arrivee_differente.keys():
                    arrivee_differente[_jour] = []
                precision = 'Si on doit venir vous chercher à la gare, précisez où et à quelle heure'
                if precision in item_dict.keys():
                    arrivee_differente[_jour].append("%s (%s) : %s" % (nom,
                                                                       item_dict['Téléphone mobile'][0],
                                                                       item_dict[precision][0]))
                else:
                    arrivee_differente[_jour].append("%s : à préciser (%s)" % (nom,
                                                                               item_dict['Téléphone mobile'][0]))
            # Transport entre les colones
            elif 'Je souhaite être véhiculé d\'une colonne à une autre' in _clef:
                sens = _clef.split(' autre ')[1]
                if sens not in transport.keys():
                    transport[sens] = {}
                for _jour in _item:
                    if _jour not in transport[sens].keys():
                        transport[sens][_jour] = []
                    try:
                        transport[sens][_jour].append("%s (%s)" % (nom, item_dict['Téléphone mobile'][0]))
                    except KeyError:
                        transport[sens][_jour].append("%s (pas de téléphone)" % nom)
            elif 'Je souhaite recevoir un plan de Chartres' in _clef:
                if 'autre' not in transport.keys():
                    transport['autre'] = {}
                if "n" not in _item[0].lower():
                    if 'plan de Chartres' not in transport['autre'].keys():
                        transport['autre']['plan de Chartres'] = []
                    try:
                        transport['autre']['plan de Chartres'].append("%s (%s - %s)" % 
                                                                      (nom,
                                                                       item_dict['Téléphone mobile'][0],
                                                                       item_dict['courriel'])
                                                                      )
                    except KeyError:
                        try:
                            transport['autre']['plan de Chartres'].append("%s (%s)" % (nom, item_dict['courriel']))
                        except KeyError:
                            transport['autre']['plan de Chartres'].append("%s (ni téléphone ni mail)" % nom)
            # Départ en cours de route
            elif 'Si je pars avant l\'arrivée à Paris, je précise quand :' in _clef:
                _jour = _clef.split(' : ')[1]
                if item_dict[_clef][0] in dico_choix['Si je pars avant l\'arrivée à Paris, je précise quand :'][0]:
                    continue
                if _jour not in transport.keys():
                    depart_different[_jour] = []
                precision = 'Si on doit vous conduire quelque part, précisez où et à quelle heure'
                if precision in item_dict.keys():
                    depart_different[_jour].append("%s (%s) : %s" % (nom,
                                                                     item_dict['Téléphone mobile'][0],
                                                                     item_dict[precision][0]))
                else:
                    depart_different[_jour].append("%s : à préciser (%s)" % (nom,
                                                                             item_dict['Téléphone mobile'][0]))
            # repas colone adulte
            elif 'Merci de me prévoir un repas au bivouac ou sur la colonne adultes' in _clef:
                _mes_choix=[]
                for _clef in dico_choix.keys():
                    if 'Merci de me prévoir un repas au bivouac ou sur la colonne adultes' in _clef:
                        _mes_choix = dico_choix[_clef]
                        break
                if len(_item) == 0:
                    _item.append('du petit déjeuner du samedi jusqu\'au déjeuner du lundi')
                    print("%s : pas de réponse pour la présence aux repas" % nom)
                for _jour in _item:
                    if 'du petit déjeuner du samedi jusqu\'au déjeuner du lundi' in _jour or \
                            'Saturday breakfast to Monday lunch' in _jour or \
                            'vom Frühstück am Samstag bis zum Mittagessen am Montag' in _jour:
                        for _repas in repas_adulte.keys():
                            repas_adulte[_repas] += 1
                        for _, _repas_nom in repas_adulte_nominatif.items():
                            _repas_nom.append(nom)
                            pass
                        break
                    elif _jour == "aucun":
                        break
                    else:
                        for list_j in _mes_choix:
                            if _jour in list_j:
                                repas_adulte[list_j[0]] += 1
                                repas_adulte_nominatif[list_j[0]].append(nom)
                                break
            # repas colone enfant
            elif 'Si je me trouve à midi sur la colonne enfants je m\'inscris pour être prévu au repas enfant' in _clef:
                _mes_choix=[]
                for _clef in dico_choix.keys():
                    if 'Si je me trouve à midi sur la colonne enfants je m\'inscris pour être prévu au repas enfant' in _clef:
                        _mes_choix = dico_choix[_clef]
                        break
                for _jour in _item:
                    for list_j in _mes_choix:
                        if _jour in list_j:
                            repas_enfant[list_j[0]] += 1
                            repas_enfant_nominatif[list_j[0]].append(nom)
                            break
            # Lit dans la tente prêtres (clergé)
            elif 'Je souhaite avoir un lit dans la tente prêtres (clergé)' in _clef:
                if _item[0] not in ['NON', 'NO', 'NEIN']:
                    divers['lit en tente prêtres'] += 1
            elif 'Chapitre des porteurs de croix (tête de colonne)' in _clef:
                for _jour in _item:
                    for demis in dico_choix["J'accepte d'aider les services suivants :"]:
                        if _jour in demis:
                            porteurs_de_croix[demis[0]].append(nom)
            elif "Service 'Régulation sécurité' (marche avec le jalon)" in _clef:
                for _jour in _item:
                    for demis in dico_choix["J'accepte d'aider les services suivants :"]:
                        if _jour in demis:
                            regulation_securite[demis[0]].append(nom)
            elif "Aide aux chapitres en manque d'aumônier" in _clef:
                for _jour in _item:
                    for demis in dico_choix["J'accepte d'aider les services suivants :"]:
                        if _jour in demis:
                            aides_chapitres[demis[0]].append(nom)
            elif "Logistique des bivouacs (par demi-journée" in _clef:
                for _jour in _item:
                    for demis in dico_choix["J'accepte d'aider les services suivants :"]:
                        if _jour in demis:
                            logistique_bivouacs[demis[0]].append(nom)
            elif "Logistique des haltes (service d'1/2h" in _clef:
                for _jour in _item:
                    for demis in dico_choix["J'accepte d'aider les services suivants :"]:
                        if _jour in demis:
                            logistique_haltes[demis[0]].append(nom)
            elif "Éclopés : méditations, chants et confessions" in _clef:
                for _jour in _item:
                    for demis in dico_choix["J'accepte d'aider les services suivants :"]:
                        if _jour in demis:
                            eclopes[demis[0]].append(nom)
            elif "Je célébrerai la sainte messe Bivouac Greffiers" in _clef:
                for _jour in _item:
                    for jours in dico_choix['Je célébrerai la sainte messe']:
                        if _jour in jours:
                            messes_greffiers[jours[0]].append(nom)
            elif "Je célébrerai la sainte messe Bivouac Trou Moreau" in _clef:
                for _jour in _item:
                    for jours in dico_choix['Je célébrerai la sainte messe']:
                        if _jour in jours:
                            messes_trou_moreau[jours[0]].append(nom)
            elif "Je célébrerai la sainte messe St Nicolas du Chardonnet" in _clef:
                for _jour in _item:
                    for jours in dico_choix['Je célébrerai la sainte messe']:
                        if _jour in jours:
                            messes_st_nic[jours[0]].append(nom)
            elif "Je célébrerai la sainte messe Chartres" in _clef:
                for _jour in _item:
                    for jours in dico_choix['Je célébrerai la sainte messe']:
                        if _jour in jours:
                            messes_chartres[jours[0]].append(nom)
            elif "J'accepte de célébrer la messe pour les services de la logistique aux horaires suivants :" in _clef:
                for _jour in _item:
                    try:
                        messes_logistiques[_jour.split(' (')[0]].append(nom)
                    except KeyError:
                        if 'Option 7' in _jour.split(' (')[0]:
                            pass  # ligne supplémentaire dans le questionnaire allemand
                        else:
                            print('%s %s' % (nom, _item))
            elif "Observations ou souhaits particuliers" in _clef:
                for _jour in _item:
                    divers.update({nom: _item[0]})
            # else:
            #     print(_clef)
    #
    # TODO ajouter les repas de midi pour les transports entre les colonnes + mail si E+A
    for _sens, transport_i in transport.items():
        for _jour, transport_j in transport_i.items():
            for indice, personne in enumerate(transport_j):
                _nom = personne.split(" (")[0]
                if 'samedi' in _jour:
                    nb_repas = 0
                    if _nom in repas_adulte_nominatif['samedi - déjeuner (pique-nique)']:
                        nb_repas+=1
                        transport_j[indice] += " repas A"
                    if _nom in repas_enfant_nominatif['samedi - déjeuner (pique-nique)']:
                        nb_repas+=1
                        transport_j[indice] += " repas E"
                    if nb_repas == 2:
                        transport_j[indice] += " (%s)" % dico_nom_contacts[_nom]
                elif 'dimanche' in _jour:
                    nb_repas = 0
                    if _nom in repas_adulte_nominatif['dimanche - déjeuner (pique-nique)']:
                        nb_repas+=1
                        transport_j[indice] += " repas A"
                    if _nom in repas_enfant_nominatif['dimanche - déjeuner (pique-nique)']:
                        nb_repas+=1
                        transport_j[indice] += " repas E"
                    if nb_repas == 2:
                        transport_j[indice] += " (%s)" % dico_nom_contacts[_nom]
    #
    # résultats
    wb = Workbook()
    wb2 = Workbook()

    # grab the active worksheet
    vers_xlsx(liste_inscrits, 'liste des inscrits', wb, 'B2')
    vers_xlsx(arrivee_differente, 'arrivée différente', wb, 'A2')
    vers_xlsx(depart_different, 'départ différent', wb, 'A2')
    vers_xlsx(transport, 'transport entre colones', wb, 'A3')
    vers_xlsx(repas_adulte, 'repas colone adulte', wb)
    vers_xlsx(repas_enfant, 'repas colone enfant', wb)
    vers_xlsx(divers, 'Divers', wb)
    vers_xlsx(porteurs_de_croix, 'porteurs de croix', wb, 'A2')
    vers_xlsx(regulation_securite, 'minist. régulation sécurité', wb, 'A2')
    vers_xlsx(aides_chapitres, 'minist. aux chapitres', wb, 'A2')
    vers_xlsx(logistique_bivouacs, 'minist. logistique des bivouacs', wb, 'A2')
    vers_xlsx(logistique_haltes, 'minist. logistique des haltes', wb, 'A2')
    vers_xlsx(eclopes, 'minist. éclopés', wb, 'A2')
    vers_xlsx(messes_logistiques, 'messes services logistiques', wb, 'A2')
    vers_xlsx(messes_chartres, 'messe à chartres', wb, 'A2')
    vers_xlsx(messes_st_nic, 'messe à st Nic', wb, 'A2')
    vers_xlsx(messes_trou_moreau, 'messe au Trou Moreau', wb, 'A2')
    vers_xlsx(messes_greffiers, 'messe aux Greffiers', wb, 'A2')
    vers_xlsx(repas_adulte_nominatif, 'repas adulte nominatif', wb2, 'A2')
    vers_xlsx(repas_enfant_nominatif, 'repas enfant nominatif', wb2, 'A2')

    #
    del wb['Sheet']
    del wb2['Sheet']
    # Save the file
    if os.path.exists("vieux_resultats.xlsx"):
        os.remove("vieux_resultats.xlsx")
    shutil.copyfile("resultats.xlsx", "vieux_resultats.xlsx")
    wb.save("resultats.xlsx")
    wb2.save("repas_nom.xlsx")
    comparaison_xslx("resultats.xlsx", "vieux_resultats.xlsx")
    subprocess.run(['open', "resultats.xlsx"], check=False)
